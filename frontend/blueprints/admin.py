# Admin blueprint - Administrative functions
import json
import io
from collections import defaultdict
from datetime import datetime
from flask import Blueprint, render_template, request, redirect, url_for, flash, session, make_response
from openpyxl import Workbook

from utils.supabase_utils import get_admin_client
from blueprints.helpers import get_user_id, is_admin, is_dev, get_user_role, get_habilitar_abas, registrar_historico

admin_bp = Blueprint('admin', __name__)


def validate_admin_token():
    """Valida o token JWT do Supabase Auth para operações admin"""
    user_id = get_user_id()
    token = session.get("supabase_token")
    
    if not token:
        flash("Sessão inválida. Faça login novamente.", "danger")
        return None, redirect(url_for("auth.login"))
    
    try:
        from utils.supabase_utils import get_supabase_client
        supabase = get_supabase_client()
        user_response = supabase.auth.get_user(token)
        if not user_response or not user_response.user:
            flash("Token inválido ou expirado.", "danger")
            return None, redirect(url_for("auth.login"))
    except Exception:
        flash("Token inválido.", "danger")
        return None, redirect(url_for("auth.login"))
    
    return user_id, None


@admin_bp.route("/admin")
def panel():
    if not is_admin():
        flash("Acesso restrito a administradores.", "danger")
        return redirect(url_for("dashboard"))
    
    user_id, error = validate_admin_token()
    if error:
        return error
    
    client = get_admin_client()
    users_response = client.table("perfil").select("*").execute()
    users = users_response.data or []
    
    if users:
        user_ids = [u.get("id") for u in users]
        
        cil_response = client.table("cilindro").select("user_id,id", count="exact").in_("user_id", user_ids).execute()
        ele_response = client.table("elemento").select("user_id,id", count="exact").in_("user_id", user_ids).execute()
        leitura_response = client.table("leitura").select("user_id,id", count="exact").in_("user_id", user_ids).execute()
        pre_response = client.table("pressao").select("user_id,id", count="exact").in_("user_id", user_ids).execute()
        amostra_response = client.table("amostra").select("user_id,id", count="exact").in_("user_id", user_ids).execute()
        
        cil_counts = {}
        for r in cil_response.data or []:
            cil_counts[r["user_id"]] = r.get("count", 0)
        
        ele_counts = {}
        for r in ele_response.data or []:
            ele_counts[r["user_id"]] = r.get("count", 0)
        
        leitura_counts = {}
        for r in leitura_response.data or []:
            leitura_counts[r["user_id"]] = r.get("count", 0)
        
        pre_counts = {}
        for r in pre_response.data or []:
            pre_counts[r["user_id"]] = r.get("count", 0)
        
        amostra_counts = {}
        for r in amostra_response.data or []:
            amostra_counts[r["user_id"]] = r.get("count", 0)
    
    for user in users:
        uid = user.get("id")
        
        user["nome"] = user.get("nome") or user.get("email") or uid
        user["cilindros"] = cil_counts.get(uid, 0)
        user["elementos"] = ele_counts.get(uid, 0)
        user["leituras"] = leitura_counts.get(uid, 0)
        user["pressoes"] = pre_counts.get(uid, 0)
        user["amostras"] = amostra_counts.get(uid, 0)
        if user.get("role") in ("admin", "dev"):
            user["habilitar_abas"] = {"cilindro": True, "pressao": True, "elemento": True, "leitura": True, "amostra": True, "historico": True}
        else:
            user["habilitar_abas"] = get_habilitar_abas(user["id"])
    
    users = sorted(users, key=lambda x: (x.get("nome") or x.get("email") or "").lower())
    
    return render_template("admin.html", users=users)


@admin_bp.route("/admin/toggle-user", methods=["POST"])
def toggle_user():
    if not is_admin():
        flash("Acesso restrito a administradores.", "danger")
        return redirect(url_for("dashboard"))
    
    user_id, error = validate_admin_token()
    if error:
        return error
    
    target_user_id = request.form.get("user_id")
    ativo = request.form.get("ativo") == "true"
    
    if target_user_id == get_user_id():
        flash("Você não pode desativar seu próprio usuário.", "warning")
        return redirect(url_for("admin.panel"))
    
    client = get_admin_client()
    client.table("perfil").update({"ativo": ativo}).eq("id", target_user_id).execute()
    
    invalidate_user_caches(target_user_id)
    
    acao = "ativado" if ativo else "desativado"
    registrar_historico("perfil", "atualizado", f"Usuário {acao}", get_user_id())
    
    flash(f"Usuário {'ativado' if ativo else 'desativado'} com sucesso!", "success")
    
    return redirect(url_for("admin.panel"))


@admin_bp.route("/admin/set-role", methods=["POST"])
def set_role():
    if not is_admin():
        flash("Acesso restrito a administradores.", "danger")
        return redirect(url_for("dashboard"))
    
    user_id, error = validate_admin_token()
    if error:
        return error
    
    target_user_id = request.form.get("user_id")
    role = request.form.get("role")
    
    if role not in ("admin", "usuario"):
        flash("Função inválida.", "danger")
        return redirect(url_for("admin.panel"))
    
    target_is_self = target_user_id == get_user_id()
    
    client = get_admin_client()
    target_perfil = client.table("perfil").select("role").eq("id", target_user_id).execute()
    if not target_perfil.data:
        flash("Usuário não encontrado.", "danger")
        return redirect(url_for("admin.panel"))
    
    if not is_dev():
        if target_is_self:
            if role == "usuario":
                flash("Você se rebaixou para usuário. Esta ação não pode ser desfeita por você.", "warning")
            else:
                flash("Você alterou sua própria função.", "warning")
        else:
            target_name = target_perfil.data[0].get("role", "?")
            flash(f"Atenção: você alterou a role de {target_user_id} de '{target_perfil.data[0].get('role', '?')}' para '{role}'.", "warning")
    
    client.table("perfil").update({"role": role}).eq("id", target_user_id).execute()
    
    if target_is_self:
        session.pop("cached_user_info", None)
    invalidate_user_caches(target_user_id)
    
    registrar_historico("perfil", "atualizado", f"Role alterada para {role}", get_user_id())
    
    flash(f"Função do usuário alterada para {role}!", "success")
    
    return redirect(url_for("admin.panel"))


@admin_bp.route("/admin/delete-user", methods=["POST"])
def delete_user():
    if not is_dev():
        flash("Acesso restrito a desenvolvedores.", "danger")
        return redirect(url_for("dashboard"))
    
    user_id, error = validate_admin_token()
    if error:
        return error
    
    target_user_id = request.form.get("user_id")
    
    if target_user_id == get_user_id():
        flash("Você não pode excluir seu próprio usuário.", "warning")
        return redirect(url_for("admin.panel"))
    
    client = get_admin_client()
    target_perfil = client.table("perfil").select("role").eq("id", target_user_id).execute()
    if target_perfil.data and target_perfil.data[0].get("role") == "dev":
        flash("Não é possível deletar outro desenvolvedor.", "danger")
        return redirect(url_for("admin.panel"))
    
    client.table("cilindro").delete().eq("user_id", target_user_id).execute()
    client.table("elemento").delete().eq("user_id", target_user_id).execute()
    client.table("leitura").delete().eq("user_id", target_user_id).execute()
    client.table("pressao").delete().eq("user_id", target_user_id).execute()
    client.table("historico_log").delete().eq("user_id", target_user_id).execute()
    client.table("perfil").delete().eq("id", target_user_id).execute()
    
    flash("Usuário e todos os seus dados foram excluídos!", "success")
    
    return redirect(url_for("admin.panel"))


@admin_bp.route("/admin/update-habilitar-abas", methods=["POST"])
def update_habilitar_abas():
    if not is_admin():
        flash("Acesso restrito a administradores.", "danger")
        return redirect(url_for("dashboard"))
    
    user_id, error = validate_admin_token()
    if error:
        return error
    
    target_user_id = request.form.get("user_id")
    aba = request.form.get("aba")
    habilitar = request.form.get("habilitar") == "true"
    
    if not target_user_id or not aba:
        flash("Parâmetros inválidos.", "danger")
        return redirect(url_for("admin.panel"))
    
    if aba not in ["cilindro", "pressao", "elemento", "leitura", "amostra", "historico"]:
        flash("Aba inválida.", "danger")
        return redirect(url_for("admin.panel"))
    
    client = get_admin_client()
    
    habilitar_abas = {"cilindro": False, "pressao": False, "elemento": False, "leitura": False, "amostra": False, "historico": False}
    perfil = client.table("perfil").select("habilitar_abas").eq("id", target_user_id).execute()
    if perfil.data and perfil.data[0].get("habilitar_abas"):
        habilitar_abas = perfil.data[0].get("habilitar_abas")
    
    habilitar_abas[aba] = habilitar
    
    client.table("perfil").update({"habilitar_abas": habilitar_abas}).eq("id", target_user_id).execute()
    
    # Registrar alteração de permissões no histórico
    acao = "habilitada" if habilitar else "desabilitada"
    nome_aba = {"cilindro": "Cilindros", "pressao": "Pressão", "elemento": "Elementos", "leitura": "Leitura", "historico": "Histórico"}.get(aba, aba)
    registrar_historico("perfil", "atualizado", f"Aba {nome_aba} {acao}", get_user_id())
    
    flash(f"Aba {nome_aba} {acao} com sucesso!", "success")
    
    return redirect(url_for("admin.panel"))


@admin_bp.route("/admin/user-data/<target_user_id>", methods=["GET"])
def user_data(target_user_id):
    if not is_admin():
        flash("Acesso restrito a administradores.", "danger")
        return redirect(url_for("dashboard"))
    
    user_id, error = validate_admin_token()
    if error:
        return error
    
    client = get_admin_client()
    
    page = int(request.args.get("page", 1))
    per_page = 20
    
    cilindro_total = client.table("cilindro").select("*", count="exact").eq("user_id", target_user_id).execute().count or 0
    elementos_total = client.table("elemento").select("*", count="exact").eq("user_id", target_user_id).execute().count or 0
    leituras_total = client.table("leitura").select("*", count="exact").eq("user_id", target_user_id).execute().count or 0
    pressoes_total = client.table("pressao").select("*", count="exact").eq("user_id", target_user_id).execute().count or 0
    amostras_total = client.table("amostra").select("*", count="exact").eq("user_id", target_user_id).execute().count or 0
    
    historico_offset = (page - 1) * per_page
    historico_log = client.table("historico_log").select(
        "tipo, acao, nome, created_at"
    ).eq("user_id", target_user_id).order("created_at", desc=True).range(historico_offset, historico_offset + per_page - 1).execute().data or []
    
    historico_total = client.table("historico_log").select("*", count="exact").eq("user_id", target_user_id).execute().count or 0
    
    history = [{
        "tipo": h.get("tipo"),
        "acao": h.get("acao"),
        "nome": h.get("nome"),
        "data": h.get("created_at")
    } for h in historico_log]
    
    perfil = client.table("perfil").select("*").eq("id", target_user_id).execute().data
    target_user = perfil[0] if perfil else {"id": target_user_id, "role": "unknown"}
    
    habilitar_abas = get_habilitar_abas(target_user_id) if target_user.get("role") not in ("admin", "dev") else {"cilindro": True, "pressao": True, "elemento": True, "leitura": True, "amostra": True, "historico": True}
    
    return render_template(
        "admin_user_data.html",
        target_user=target_user,
        cilindro_total=cilindro_total,
        elementos_total=elementos_total,
        leituras_total=leituras_total,
        pressoes_total=pressoes_total,
        amostras_total=amostras_total,
        habilitar_abas=habilitar_abas,
        history=history,
        historico_total=historico_total,
        page=page,
        per_page=per_page
    )


def _compute_kpis_export(leituras_data, cilindro_data, elementos_data, pressoes_data, amostras_data):
    from collections import defaultdict
    elemento_dict = {e["id"]: e for e in elementos_data}

    def parse_tempo(t):
        if not t:
            return 0
        p = t.split(":")
        try:
            if len(p) == 3:
                return int(p[0]) * 60 + int(p[1]) + int(p[2]) / 60
            if len(p) == 2:
                return int(p[0]) * 60 + int(p[1])
        except (ValueError, IndexError):
            pass
        return 0

    total_qtd = sum(a.get("quantidade", 1) for a in leituras_data)
    cilindros_ativos = sum(1 for c in cilindro_data if c.get("status") == "ativo")

    gas_por_cilindro = defaultdict(float)
    total_gas = 0.0
    for a in leituras_data:
        cid, eid = a.get("cilindro_id"), a.get("elemento_id")
        if cid and eid:
            mins = parse_tempo(a.get("tempo_chama"))
            cons = float(elemento_dict.get(eid, {}).get("consumo_lpm", 0))
            g = mins / 60 * cons
            gas_por_cilindro[cid] += g
            total_gas += g

    gas_restante = 0.0
    for c in cilindro_data:
        if c.get("status") == "ativo":
            rest = float(c.get("litros_equivalentes", 0)) - gas_por_cilindro.get(c["id"], 0)
            if rest > 0:
                gas_restante += rest

    custo_total = sum(float(c.get("custo", 0)) for c in cilindro_data if c.get("custo"))
    custo_leitura = round(custo_total / total_qtd, 2) if total_qtd > 0 else 0

    return {
        "cilindros_ativos": cilindros_ativos,
        "gas_restante_litros": round(gas_restante, 1),
        "total_leituras_quantidade": total_qtd,
        "total_amostras": len(amostras_data),
        "custo_medio_por_leitura": custo_leitura,
        "gas_consumido_litros": round(total_gas, 1),
    }


@admin_bp.route("/admin/export")
def export_data():
    if not is_admin():
        flash("Acesso restrito a administradores.", "danger")
        return redirect(url_for("dashboard"))
    
    user_id, error = validate_admin_token()
    if error:
        return error
    
    formato = request.args.get("formato", "json").lower()
    
    if formato not in ["csv", "json", "excel", "md"]:
        flash("Formato inválido.", "danger")
        return redirect(url_for("dashboard"))
    
    client = get_admin_client()
    
    cilindro_data = client.table("cilindro").select("*").execute().data or []
    elementos_data = client.table("elemento").select("*").execute().data or []
    leituras_data = client.table("leitura").select("*").execute().data or []
    pressoes_data = client.table("pressao").select("*").execute().data or []

    try:
        amostras_data = client.table("amostra").select("*").execute().data or []
    except Exception:
        amostras_data = []

    try:
        ae_data = client.table("amostra_elemento").select("*").execute().data or []
    except Exception:
        ae_data = []
    
    usuarios_data = client.table("perfil").select("id,email,nome").execute().data or []
    usuarios_dict = {u.get("id"): u for u in usuarios_data}

    for c in cilindro_data:
        uid = c.get("user_id")
        if uid:
            u = usuarios_dict.get(uid, {})
            c["usuario_email"] = u.get("email", "")
            c["usuario_nome"] = u.get("nome", "")
    
    for e in elementos_data:
        uid = e.get("user_id")
        if uid:
            u = usuarios_dict.get(uid, {})
            e["usuario_email"] = u.get("email", "")
            e["usuario_nome"] = u.get("nome", "")
    
    for p in pressoes_data:
        uid = p.get("user_id")
        if uid:
            u = usuarios_dict.get(uid, {})
            p["usuario_email"] = u.get("email", "")
            p["usuario_nome"] = u.get("nome", "")
    
    for a in amostras_data:
        uid = a.get("user_id")
        if uid:
            u = usuarios_dict.get(uid, {})
            a["usuario_email"] = u.get("email", "")
            a["usuario_nome"] = u.get("nome", "")
    
    cilindro_dict = {c.get("id"): c.get("codigo") for c in cilindro_data}
    elemento_dict = {e.get("id"): e.get("nome") for e in elementos_data}
    
    for a in leituras_data:
        uid = a.get("user_id")
        if uid:
            u = usuarios_dict.get(uid, {})
            a["usuario_email"] = u.get("email", "")
            a["usuario_nome"] = u.get("nome", "")
        a["cilindro_codigo"] = cilindro_dict.get(a.get("cilindro_id"), "")
        a["elemento_nome"] = elemento_dict.get(a.get("elemento_id"), "")
    
    for p in pressoes_data:
        p["cilindro_codigo"] = cilindro_dict.get(p.get("cilindro_id"), "")
    
    # Agrupa elementos por amostra
    amostra_e_ids = defaultdict(list)
    for ae in ae_data:
        amostra_e_ids[ae.get("amostra_id")].append(ae.get("elemento_id"))
    
    for a in amostras_data:
        eids = amostra_e_ids.get(a.get("id"), [])
        a["elementos_nomes"] = ", ".join(elemento_dict.get(eid, str(eid)) for eid in eids if eid)
        a["qtd_elementos"] = len(eids)
    
    cached = session.get('cached_user_info', {})
    user_name = (cached.get('user_name', '') or 'unknown').replace(' ', '_').replace('/', '_')
    data_str = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    filename_base = f"labgas_export_{data_str}_{user_name}"
    
    kpis = _compute_kpis_export(leituras_data, cilindro_data, elementos_data, pressoes_data, amostras_data)
    
    if formato == "json":
        data = {
            "exportado_em": datetime.now().isoformat(),
            "exportado_por": user_name,
            "kpis": kpis,
            "cilindros": cilindro_data,
            "elementos": elementos_data,
            "leituras": leituras_data,
            "pressoes": pressoes_data,
            "amostras": amostras_data,
        }
        response = make_response(json.dumps(data, indent=2, default=str))
        response.headers["Content-Disposition"] = f"attachment; filename={filename_base}.json"
        response.headers["Content-Type"] = "application/json"
        return response
    
    elif formato == "csv":
        output = io.StringIO()
        output.write(f"# Exportado por: {user_name} em {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n")
        output.write(f"# KPIs: Cilindros Ativos={kpis['cilindros_ativos']}, Gas Restante={kpis['gas_restante_litros']}L, "
                     f"Leituras={kpis['total_leituras_quantidade']}, Amostras={kpis['total_amostras']}, "
                     f"Custo/Leitura=R${kpis['custo_medio_por_leitura']}, Gas Consumido={kpis['gas_consumido_litros']}L\n\n")
        
        output.write("# CILINDROS\n")
        if cilindro_data:
            headers = ["id", "codigo", "data_compra", "data_inicio_consumo", "data_fim",
                      "gas_kg", "litros_equivalentes", "custo", "status",
                      "usuario_email", "usuario_nome", "created_at"]
            output.write(",".join(headers) + "\n")
            for row in cilindro_data:
                values = [str(row.get(h, "")) for h in headers]
                output.write(",".join(values) + "\n")
        
        output.write("\n# ELEMENTOS\n")
        if elementos_data:
            headers = ["id", "nome", "consumo_lpm", "usuario_email", "usuario_nome", "created_at"]
            output.write(",".join(headers) + "\n")
            for row in elementos_data:
                values = [str(row.get(h, "")) for h in headers]
                output.write(",".join(values) + "\n")
        
        output.write("\n# LEITURAS\n")
        if leituras_data:
            headers = ["id", "data", "tempo_chama", "cilindro_id", "cilindro_codigo",
                      "elemento_id", "elemento_nome", "quantidade",
                      "usuario_email", "usuario_nome", "created_at"]
            output.write(",".join(headers) + "\n")
            for row in leituras_data:
                values = [str(row.get(h, "")) for h in headers]
                output.write(",".join(values) + "\n")
        
        output.write("\n# PRESSOES\n")
        if pressoes_data:
            headers = ["id", "cilindro_id", "cilindro_codigo", "pressao", "temperatura", "data", "hora",
                      "usuario_email", "usuario_nome", "created_at"]
            output.write(",".join(headers) + "\n")
            for row in pressoes_data:
                values = [str(row.get(h, "")) for h in headers]
                output.write(",".join(values) + "\n")
        
        output.write("\n# AMOSTRAS\n")
        if amostras_data:
            headers = ["id", "numero_amostra", "lote", "elementos", "qtd_elementos",
                      "usuario_email", "usuario_nome", "created_at"]
            output.write(",".join(headers) + "\n")
            for row in amostras_data:
                values = [str(row.get(h, "")) for h in headers]
                output.write(",".join(values) + "\n")
        
        response = make_response(output.getvalue())
        response.headers["Content-Disposition"] = f"attachment; filename={filename_base}.csv"
        response.headers["Content-Type"] = "text/csv"
        return response
    
    elif formato == "excel":
        wb = Workbook()
        
        ws_resumo = wb.active
        ws_resumo.title = "Resumo"
        ws_resumo.append(["Indicador", "Valor"])
        for k, v in kpis.items():
            ws_resumo.append([k.replace("_", " ").title(), v])
        ws_resumo.append(["Exportado por", user_name])
        ws_resumo.append(["Exportado em", datetime.now().strftime('%d/%m/%Y %H:%M:%S')])
        
        ws_cilindros = wb.create_sheet("Cilindros")
        if cilindro_data:
            headers = ["ID", "Código", "Data Compra", "Data Início", "Data Fim",
                      "Gas (kg)", "Litros", "Custo", "Status",
                      "Usuário Email", "Usuário Nome", "Criado em"]
            ws_cilindros.append(headers)
            for row in cilindro_data:
                ws_cilindros.append([
                    row.get("id"), row.get("codigo"), row.get("data_compra"),
                    row.get("data_inicio_consumo"), row.get("data_fim"),
                    row.get("gas_kg"), row.get("litros_equivalentes"), row.get("custo"),
                    row.get("status"),
                    row.get("usuario_email"), row.get("usuario_nome"), row.get("created_at")
                ])
        
        ws_elementos = wb.create_sheet("Elementos")
        if elementos_data:
            headers = ["ID", "Nome", "Consumo (L/min)", "Usuário Email", "Usuário Nome", "Criado em"]
            ws_elementos.append(headers)
            for row in elementos_data:
                ws_elementos.append([
                    row.get("id"), row.get("nome"), row.get("consumo_lpm"),
                    row.get("usuario_email"), row.get("usuario_nome"), row.get("created_at")
                ])
        
        ws_leituras = wb.create_sheet("Leituras")
        if leituras_data:
            headers = ["ID", "Data", "Tempo Chama", "Cilindro ID", "Cilindro Código",
                      "Elemento ID", "Elemento Nome", "Qtd",
                      "Usuário Email", "Usuário Nome", "Criado em"]
            ws_leituras.append(headers)
            for row in leituras_data:
                ws_leituras.append([
                    row.get("id"), row.get("data"), row.get("tempo_chama"),
                    row.get("cilindro_id"), row.get("cilindro_codigo"),
                    row.get("elemento_id"), row.get("elemento_nome"),
                    row.get("quantidade"),
                    row.get("usuario_email"), row.get("usuario_nome"), row.get("created_at")
                ])
        
        ws_pressoes = wb.create_sheet("Pressoes")
        if pressoes_data:
            headers = ["ID", "Cilindro ID", "Cilindro Código", "Pressão (bar)", "Temperatura (°C)", "Data", "Hora",
                      "Usuário Email", "Usuário Nome", "Criado em"]
            ws_pressoes.append(headers)
            for row in pressoes_data:
                ws_pressoes.append([
                    row.get("id"), row.get("cilindro_id"), row.get("cilindro_codigo"),
                    row.get("pressao"), row.get("temperatura"), row.get("data"), row.get("hora"),
                    row.get("usuario_email"), row.get("usuario_nome"), row.get("created_at")
                ])
        
        ws_amostras = wb.create_sheet("Amostras")
        if amostras_data:
            headers = ["ID", "Número", "Lote", "Elementos", "Qtd Elementos",
                      "Usuário Email", "Usuário Nome", "Criado em"]
            ws_amostras.append(headers)
            for row in amostras_data:
                ws_amostras.append([
                    row.get("id"), row.get("numero_amostra"), row.get("lote"),
                    row.get("elementos_nomes"), row.get("qtd_elementos"),
                    row.get("usuario_email"), row.get("usuario_nome"), row.get("created_at")
                ])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers["Content-Disposition"] = f"attachment; filename={filename_base}.xlsx"
        response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        return response
    
    elif formato == "md":
        md_output = io.StringIO()
        md_output.write(f"# LabGas Manager — Exportação\n\n")
        md_output.write(f"**Exportado por:** {user_name}\n")
        md_output.write(f"**Exportado em:** {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n\n")
        md_output.write(f"## KPIs\n\n")
        md_output.write(f"| Indicador | Valor |\n|---|---|\n")
        for k, v in kpis.items():
            md_output.write(f"| {k.replace('_', ' ').title()} | {v} |\n")
        md_output.write(f"\n**Totais:** {len(cilindro_data)} Cilindros | {len(pressoes_data)} Pressões | {len(elementos_data)} Elementos | {len(leituras_data)} Leituras | {len(amostras_data)} Amostras\n\n")
        
        md_output.write("## Cilindros\n\n")
        if cilindro_data:
            md_output.write("| ID | Código | Status | Gas (kg) | Custo | Usuário |\n")
            md_output.write("|---|---|---|---|---|---|\n")
            for row in cilindro_data:
                md_output.write(f"| {row.get('id')} | {row.get('codigo')} | {row.get('status')} | {row.get('gas_kg')} | R${row.get('custo')} | {row.get('usuario_email')} |\n")
        else:
            md_output.write("*Nenhum cilindro encontrado.*\n\n")
        
        md_output.write("\n## Pressões\n\n")
        if pressoes_data:
            md_output.write("| ID | Cilindro | Pressão (bar) | Temperatura (°C) | Data | Hora | Usuário |\n")
            md_output.write("|---|---|---|---|---|---|---|\n")
            for row in pressoes_data:
                md_output.write(f"| {row.get('id')} | {row.get('cilindro_codigo')} | {row.get('pressao')} | {row.get('temperatura')} | {row.get('data')} | {row.get('hora')} | {row.get('usuario_email')} |\n")
        else:
            md_output.write("*Nenhuma pressão encontrada.*\n\n")
        
        md_output.write("\n## Elementos\n\n")
        if elementos_data:
            md_output.write("| ID | Nome | Consumo (L/min) | Usuário |\n")
            md_output.write("|---|---|---|---|\n")
            for row in elementos_data:
                md_output.write(f"| {row.get('id')} | {row.get('nome')} | {row.get('consumo_lpm')} | {row.get('usuario_email')} |\n")
        else:
            md_output.write("*Nenhum elemento encontrado.*\n\n")
        
        md_output.write("\n## Leituras\n\n")
        if leituras_data:
            md_output.write("| ID | Data | Tempo | Cilindro | Elemento | Qtd | Usuário |\n")
            md_output.write("|---|---|---|---|---|---|---|\n")
            for row in leituras_data:
                md_output.write(f"| {row.get('id')} | {row.get('data')} | {row.get('tempo_chama')} | {row.get('cilindro_codigo')} | {row.get('elemento_nome')} | {row.get('quantidade')} | {row.get('usuario_email')} |\n")
        else:
            md_output.write("*Nenhuma leitura encontrada.*\n\n")
        
        md_output.write("\n## Amostras\n\n")
        if amostras_data:
            md_output.write("| ID | Número | Lote | Elementos | Qtd | Usuário |\n")
            md_output.write("|---|---|---|---|---|---|\n")
            for row in amostras_data:
                md_output.write(f"| {row.get('id')} | {row.get('numero_amostra')} | {row.get('lote')} | {row.get('elementos_nomes')} | {row.get('qtd_elementos')} | {row.get('usuario_email')} |\n")
        else:
            md_output.write("*Nenhuma amostra encontrada.*\n\n")
        
        response = make_response(md_output.getvalue())
        response.headers["Content-Disposition"] = f"attachment; filename={filename_base}.md"
        response.headers["Content-Type"] = "text/markdown"
        return response
    
    flash("Formato não suportado.", "danger")
    return redirect(url_for("dashboard"))
